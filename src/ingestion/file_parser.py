"""
file_parser.py — Parse uploaded client files (CSV, Excel, PDF) into structured data.

Extracts tabular data and text content from common report formats.
Returns normalised dicts ready for injection into the AI pipeline or DB storage.
"""
from __future__ import annotations

import csv
import io
import logging
import os
from typing import Any, Optional

logger = logging.getLogger(__name__)


# ── Result container ──────────────────────────────────────────────────────────

class ParseResult:
    """Holds parsed data from a single file."""

    __slots__ = ("filename", "file_type", "rows", "text", "metadata", "errors")

    def __init__(
        self,
        filename: str,
        file_type: str,
        rows: list[dict[str, Any]] | None = None,
        text: str = "",
        metadata: dict[str, Any] | None = None,
        errors: list[str] | None = None,
    ) -> None:
        self.filename = filename
        self.file_type = file_type
        self.rows = rows or []
        self.text = text
        self.metadata = metadata or {}
        self.errors = errors or []

    @property
    def ok(self) -> bool:
        return len(self.errors) == 0

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "file_type": self.file_type,
            "row_count": self.row_count,
            "text_length": len(self.text),
            "rows": self.rows,
            "text": self.text,
            "metadata": self.metadata,
            "errors": self.errors,
        }

    def summary(self) -> str:
        """One-line summary for logging / injection into AI prompts."""
        if self.rows:
            cols = list(self.rows[0].keys()) if self.rows else []
            return f"{self.filename}: {self.row_count} rows, columns={cols}"
        if self.text:
            return f"{self.filename}: {len(self.text)} chars of text"
        return f"{self.filename}: empty or unparseable"


# ── CSV Parser ────────────────────────────────────────────────────────────────

def parse_csv(filepath: str) -> ParseResult:
    """Parse a CSV file into a list of row dicts."""
    filename = os.path.basename(filepath)
    try:
        with open(filepath, "r", newline="", encoding="utf-8-sig") as f:
            # Sniff delimiter
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel

            reader = csv.DictReader(f, dialect=dialect)
            rows = []
            for i, row in enumerate(reader):
                # Normalise keys: strip whitespace, lowercase
                clean = {k.strip().lower().replace(" ", "_"): _coerce_value(v) for k, v in row.items() if k}
                rows.append(clean)
                if i >= 49_999:  # safety cap
                    break

            metadata = {
                "delimiter": getattr(dialect, "delimiter", ","),
                "columns": list(rows[0].keys()) if rows else [],
                "row_count": len(rows),
            }
            logger.info("Parsed CSV %s: %d rows, %d columns", filename, len(rows), len(metadata["columns"]))
            return ParseResult(filename=filename, file_type="csv", rows=rows, metadata=metadata)

    except Exception as exc:
        logger.error("CSV parse failed for %s: %s", filepath, exc)
        return ParseResult(filename=filename, file_type="csv", errors=[str(exc)])


# ── Excel Parser ──────────────────────────────────────────────────────────────

def parse_excel(filepath: str, sheet_name: str | int = 0) -> ParseResult:
    """Parse an Excel (.xlsx) file into a list of row dicts using openpyxl."""
    filename = os.path.basename(filepath)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, read_only=True, data_only=True)
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        # First row = headers
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            wb.close()
            return ParseResult(filename=filename, file_type="excel", errors=["Empty worksheet"])

        headers = [
            (str(h).strip().lower().replace(" ", "_") if h is not None else f"col_{i}")
            for i, h in enumerate(raw_headers)
        ]

        rows: list[dict[str, Any]] = []
        for i, row_vals in enumerate(rows_iter):
            if i >= 49_999:
                break
            row_dict = {}
            for j, val in enumerate(row_vals):
                if j < len(headers):
                    row_dict[headers[j]] = _coerce_value(val)
            rows.append(row_dict)

        wb.close()

        metadata = {
            "sheet": ws.title,
            "columns": headers,
            "row_count": len(rows),
            "total_sheets": len(wb.sheetnames) if hasattr(wb, "sheetnames") else 1,
        }
        logger.info("Parsed Excel %s: %d rows, %d columns", filename, len(rows), len(headers))
        return ParseResult(filename=filename, file_type="excel", rows=rows, metadata=metadata)

    except ImportError:
        msg = "openpyxl not installed — cannot parse Excel files"
        logger.warning(msg)
        return ParseResult(filename=filename, file_type="excel", errors=[msg])
    except Exception as exc:
        logger.error("Excel parse failed for %s: %s", filepath, exc)
        return ParseResult(filename=filename, file_type="excel", errors=[str(exc)])


# ── PDF Parser ────────────────────────────────────────────────────────────────

def parse_pdf(filepath: str, max_pages: int = 50) -> ParseResult:
    """Extract text from a PDF using pdfplumber. Also attempts table extraction."""
    filename = os.path.basename(filepath)
    try:
        import pdfplumber

        text_parts: list[str] = []
        all_table_rows: list[dict[str, Any]] = []

        with pdfplumber.open(filepath) as pdf:
            page_count = len(pdf.pages)
            for i, page in enumerate(pdf.pages[:max_pages]):
                page_text = page.extract_text() or ""
                text_parts.append(page_text)

                # Try to extract tables
                tables = page.extract_tables() or []
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    raw_headers = table[0]
                    headers = [
                        (str(h).strip().lower().replace(" ", "_") if h else f"col_{j}")
                        for j, h in enumerate(raw_headers)
                    ]
                    for data_row in table[1:]:
                        row_dict = {}
                        for j, val in enumerate(data_row):
                            if j < len(headers):
                                row_dict[headers[j]] = _coerce_value(val)
                        all_table_rows.append(row_dict)

        full_text = "\n\n".join(text_parts)
        metadata = {
            "page_count": page_count,
            "pages_parsed": min(page_count, max_pages),
            "text_length": len(full_text),
            "tables_extracted": len(all_table_rows),
        }

        logger.info(
            "Parsed PDF %s: %d pages, %d chars, %d table rows",
            filename, page_count, len(full_text), len(all_table_rows),
        )
        return ParseResult(
            filename=filename,
            file_type="pdf",
            rows=all_table_rows,
            text=full_text,
            metadata=metadata,
        )

    except ImportError:
        msg = "pdfplumber not installed — cannot parse PDF files"
        logger.warning(msg)
        return ParseResult(filename=filename, file_type="pdf", errors=[msg])
    except Exception as exc:
        logger.error("PDF parse failed for %s: %s", filepath, exc)
        return ParseResult(filename=filename, file_type="pdf", errors=[str(exc)])


# ── Auto-detect and parse ─────────────────────────────────────────────────────

_EXTENSION_MAP = {
    ".csv": parse_csv,
    ".tsv": parse_csv,
    ".xlsx": parse_excel,
    ".xls": parse_excel,
    ".pdf": parse_pdf,
}


def parse_file(filepath: str) -> ParseResult:
    """Auto-detect file type by extension and parse."""
    ext = os.path.splitext(filepath)[1].lower()
    parser = _EXTENSION_MAP.get(ext)
    if parser is None:
        return ParseResult(
            filename=os.path.basename(filepath),
            file_type="unknown",
            errors=[f"Unsupported file type: {ext}"],
        )
    return parser(filepath)


def parse_files(filepaths: list[str]) -> list[ParseResult]:
    """Parse multiple files and return all results."""
    return [parse_file(fp) for fp in filepaths]


# ── KPI Extraction ────────────────────────────────────────────────────────────

# Common KPI column name patterns
_KPI_PATTERNS: dict[str, list[str]] = {
    "impressions": ["impressions", "impr", "imp", "reach"],
    "clicks": ["clicks", "click", "clk"],
    "ctr": ["ctr", "click_through_rate", "click_rate"],
    "cpc": ["cpc", "cost_per_click", "avg_cpc"],
    "conversions": ["conversions", "conv", "purchases", "leads"],
    "cpa": ["cpa", "cost_per_acquisition", "cost_per_conversion", "cost/conv"],
    "roas": ["roas", "return_on_ad_spend"],
    "spend": ["spend", "cost", "ad_spend", "budget_spent", "amount_spent"],
    "revenue": ["revenue", "sales", "income", "doanh_thu"],
}


def extract_kpis_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Extract and aggregate KPI values from parsed tabular data.
    Returns a normalised KPI dict ready for injection into task context.
    """
    if not rows:
        return {}

    # Map actual column names to canonical KPI names
    sample_keys = set()
    for row in rows[:5]:
        sample_keys.update(row.keys())

    col_mapping: dict[str, str] = {}  # canonical -> actual column name
    for canonical, patterns in _KPI_PATTERNS.items():
        for actual_col in sample_keys:
            if actual_col.lower() in patterns or any(p in actual_col.lower() for p in patterns):
                col_mapping[canonical] = actual_col
                break

    if not col_mapping:
        return {}

    # Aggregate: sum for volume metrics, average for rate metrics
    _SUM_METRICS = {"impressions", "clicks", "conversions", "spend", "revenue"}
    _AVG_METRICS = {"ctr", "cpc", "cpa", "roas"}

    kpis: dict[str, Any] = {}
    for canonical, col in col_mapping.items():
        values = []
        for row in rows:
            val = row.get(col)
            if val is not None:
                numeric = _to_float(val)
                if numeric is not None:
                    values.append(numeric)

        if not values:
            continue

        if canonical in _SUM_METRICS:
            kpis[canonical] = round(sum(values), 2)
        elif canonical in _AVG_METRICS:
            kpis[canonical] = round(sum(values) / len(values), 4)
        else:
            kpis[canonical] = round(sum(values) / len(values), 2)

    # Derive calculated metrics if raw data allows
    if "spend" in kpis and "clicks" in kpis and kpis["clicks"] > 0:
        kpis.setdefault("cpc", round(kpis["spend"] / kpis["clicks"], 2))
    if "impressions" in kpis and "clicks" in kpis and kpis["impressions"] > 0:
        kpis.setdefault("ctr", round(kpis["clicks"] / kpis["impressions"] * 100, 2))
    if "spend" in kpis and "conversions" in kpis and kpis["conversions"] > 0:
        kpis.setdefault("cpa", round(kpis["spend"] / kpis["conversions"], 2))
    if "revenue" in kpis and "spend" in kpis and kpis["spend"] > 0:
        kpis.setdefault("roas", round(kpis["revenue"] / kpis["spend"], 2))

    return kpis


# ── Helpers ───────────────────────────────────────────────────────────────────

def _coerce_value(val: Any) -> Any:
    """Try to convert string values to numbers where appropriate."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "").replace("%", "").replace("$", "").replace("₫", "")
        if not cleaned:
            return val.strip()
        try:
            if "." in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return val.strip()
    return val


def _to_float(val: Any) -> Optional[float]:
    """Convert a value to float, returning None if impossible."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "").replace("%", "").replace("$", "").replace("₫", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None
